import os
import re
import tempfile
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED

from lxml import etree
from openpyxl import load_workbook
from docx2pdf import convert


NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def limpar_nome_arquivo(nome):
    nome = str(nome).strip()
    nome = re.sub(r'[\\/:*?"<>|]', '', nome)
    nome = re.sub(r'\s+', ' ', nome)
    return nome


def formatar_valor(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def indice_para_letra(indice_zero_based):
    resultado = ""
    indice = indice_zero_based + 1

    while indice > 0:
        indice, resto = divmod(indice - 1, 26)
        resultado = chr(65 + resto) + resultado

    return resultado


def letra_para_indice(letra_coluna):
    letra_coluna = letra_coluna.strip().upper()
    resultado = 0

    for char in letra_coluna:
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Coluna inválida: {letra_coluna}")
        resultado = resultado * 26 + (ord(char) - ord("A") + 1)

    return resultado - 1


def remover_marcacao_amarela(elemento):
    for rPr in elemento.xpath(".//w:rPr", namespaces=NS):
        for hl in rPr.xpath("./w:highlight", namespaces=NS):
            rPr.remove(hl)
        for shd in rPr.xpath("./w:shd", namespaces=NS):
            rPr.remove(shd)


def preencher_content_controls(arquivo_entrada, arquivo_saida, valores):
    with tempfile.TemporaryDirectory() as pasta_temp:
        with ZipFile(arquivo_entrada, "r") as docx:
            docx.extractall(pasta_temp)

        caminho_xml = os.path.join(pasta_temp, "word", "document.xml")
        tree = etree.parse(caminho_xml)
        root = tree.getroot()

        for sdt in root.xpath(".//w:sdt", namespaces=NS):
            tag = sdt.xpath("./w:sdtPr/w:tag/@w:val", namespaces=NS)
            alias = sdt.xpath("./w:sdtPr/w:alias/@w:val", namespaces=NS)

            identificador = None

            if tag and tag[0] in valores:
                identificador = tag[0]
            elif alias and alias[0] in valores:
                identificador = alias[0]

            if not identificador:
                continue

            novo_valor = valores[identificador]
            sdt_content = sdt.find("w:sdtContent", namespaces=NS)

            if sdt_content is None:
                continue

            remover_marcacao_amarela(sdt_content)
            textos = sdt_content.xpath(".//w:t", namespaces=NS)

            if textos:
                textos[0].text = novo_valor
                for t in textos[1:]:
                    t.text = ""
            else:
                p = sdt_content.find("w:p", namespaces=NS)
                if p is None:
                    p = etree.SubElement(
                        sdt_content,
                        "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"
                    )

                r = etree.SubElement(
                    p,
                    "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}r"
                )

                etree.SubElement(
                    r,
                    "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}rPr"
                )

                t = etree.SubElement(
                    r,
                    "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
                )
                t.text = novo_valor

        tree.write(
            caminho_xml,
            xml_declaration=True,
            encoding="UTF-8",
            standalone="yes"
        )

        with ZipFile(arquivo_saida, "w", ZIP_DEFLATED) as novo_docx:
            for pasta_raiz, _, arquivos in os.walk(pasta_temp):
                for arquivo in arquivos:
                    caminho_completo = os.path.join(pasta_raiz, arquivo)
                    caminho_relativo = os.path.relpath(caminho_completo, pasta_temp)
                    novo_docx.write(caminho_completo, caminho_relativo)


def gerar_nome_unico(caminho_arquivo):
    if not os.path.exists(caminho_arquivo):
        return caminho_arquivo

    base, ext = os.path.splitext(caminho_arquivo)
    contador = 1

    while True:
        novo_caminho = f"{base}_{contador}{ext}"
        if not os.path.exists(novo_caminho):
            return novo_caminho
        contador += 1


def obter_primeiro_arquivo(pasta, extensoes):
    if not os.path.exists(pasta):
        raise FileNotFoundError(f"A pasta não existe: {pasta}")

    arquivos = []
    for nome in os.listdir(pasta):
        caminho_completo = os.path.join(pasta, nome)
        if os.path.isfile(caminho_completo):
            if any(nome.lower().endswith(ext.lower()) for ext in extensoes):
                arquivos.append(caminho_completo)

    arquivos.sort()

    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum arquivo com extensão {extensoes} foi encontrado na pasta: {pasta}"
        )

    return arquivos[0]


def obter_todos_arquivos(pasta, extensoes):
    if not os.path.exists(pasta):
        raise FileNotFoundError(f"A pasta não existe: {pasta}")

    arquivos = []
    for nome in os.listdir(pasta):
        caminho_completo = os.path.join(pasta, nome)
        if os.path.isfile(caminho_completo):
            if any(nome.lower().endswith(ext.lower()) for ext in extensoes):
                arquivos.append(caminho_completo)

    arquivos.sort()

    if not arquivos:
        raise FileNotFoundError(
            f"Nenhum arquivo com extensão {extensoes} foi encontrado na pasta: {pasta}"
        )

    return arquivos


def gerar_mapeamento_automatico(quantidade_campos):
    mapeamento = {}
    for i in range(1, quantidade_campos + 1):
        mapeamento[str(i)] = indice_para_letra(i - 1)
    return mapeamento


def contar_campos_word(arquivo_word):
    with tempfile.TemporaryDirectory() as pasta_temp:
        with ZipFile(arquivo_word, "r") as docx:
            docx.extractall(pasta_temp)

        caminho_xml = os.path.join(pasta_temp, "word", "document.xml")
        tree = etree.parse(caminho_xml)
        root = tree.getroot()

        identificadores = set()

        for sdt in root.xpath(".//w:sdt", namespaces=NS):
            tag = sdt.xpath("./w:sdtPr/w:tag/@w:val", namespaces=NS)
            alias = sdt.xpath("./w:sdtPr/w:alias/@w:val", namespaces=NS)

            if tag and str(tag[0]).strip():
                identificadores.add(str(tag[0]).strip())
            elif alias and str(alias[0]).strip():
                identificadores.add(str(alias[0]).strip())

        return {
            "total_campos": len(identificadores),
            "campos": sorted(identificadores),
        }


def contar_colunas_excel(arquivo_excel):
    wb = load_workbook(arquivo_excel, data_only=True)
    ws = wb.active

    primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    total_colunas = 0
    for valor in primeira_linha:
        if valor is not None and str(valor).strip():
            total_colunas += 1

    return {
        "total_colunas": total_colunas
    }


def criar_estrutura_pastas_modelo(pasta_raiz):
    pasta_entrada = os.path.join(pasta_raiz, "entrada")
    pasta_base = os.path.join(pasta_raiz, "base")
    pasta_saida = os.path.join(pasta_raiz, "saida")
    pasta_word = os.path.join(pasta_saida, "word")
    pasta_pdf = os.path.join(pasta_saida, "pdf")

    caminhos = [pasta_entrada, pasta_base, pasta_saida, pasta_word, pasta_pdf]

    criou_alguma = False
    for caminho in caminhos:
        if not os.path.exists(caminho):
            os.makedirs(caminho, exist_ok=True)
            criou_alguma = True

    return {
        "entrada": pasta_entrada,
        "base": pasta_base,
        "saida": pasta_saida,
        "word": pasta_word,
        "pdf": pasta_pdf,
        "criou_alguma": criou_alguma,
    }


def criar_estrutura_pastas_conversao(pasta_raiz):
    pasta_base = os.path.join(pasta_raiz, "base")
    pasta_saida = os.path.join(pasta_raiz, "saida")
    pasta_pdf = os.path.join(pasta_saida, "pdf")

    caminhos = [pasta_base, pasta_saida, pasta_pdf]

    criou_alguma = False
    for caminho in caminhos:
        if not os.path.exists(caminho):
            os.makedirs(caminho, exist_ok=True)
            criou_alguma = True

    return {
        "base": pasta_base,
        "saida": pasta_saida,
        "pdf": pasta_pdf,
        "criou_alguma": criou_alguma,
    }


def localizar_arquivos_modelo(pasta_raiz):
    estrutura = criar_estrutura_pastas_modelo(pasta_raiz)

    arquivo_excel = obter_primeiro_arquivo(
        estrutura["entrada"],
        extensoes=[".xlsx", ".xlsm", ".xltx", ".xltm"]
    )

    arquivo_word = obter_primeiro_arquivo(
        estrutura["base"],
        extensoes=[".docx"]
    )

    return {
        "arquivo_excel": arquivo_excel,
        "arquivo_word": arquivo_word,
        "estrutura": estrutura,
    }


def localizar_arquivos_conversao(pasta_raiz):
    estrutura = criar_estrutura_pastas_conversao(pasta_raiz)

    arquivo_word = obter_primeiro_arquivo(
        estrutura["base"],
        extensoes=[".docx"]
    )

    return {
        "arquivo_word": arquivo_word,
        "estrutura": estrutura,
    }


def executar_processamento_modelo(pasta_raiz, quantidade_campos, gerar_pdf=True):
    estrutura = criar_estrutura_pastas_modelo(pasta_raiz)

    arquivo_excel = obter_primeiro_arquivo(
        estrutura["entrada"],
        extensoes=[".xlsx", ".xlsm", ".xltx", ".xltm"]
    )

    arquivo_word_base = obter_primeiro_arquivo(
        estrutura["base"],
        extensoes=[".docx"]
    )

    wb = load_workbook(arquivo_excel, data_only=True)
    ws = wb.active

    mapeamento_campos = gerar_mapeamento_automatico(quantidade_campos)
    indice_nome = letra_para_indice("A")

    total_word = 0
    total_pdf = 0

    for linha in ws.iter_rows(min_row=2, values_only=True):
        nome_arquivo = formatar_valor(linha[indice_nome]) if len(linha) > indice_nome else ""

        if not nome_arquivo:
            continue

        campos = {}

        for identificador, letra_coluna in mapeamento_campos.items():
            indice_coluna = letra_para_indice(letra_coluna)
            valor = formatar_valor(linha[indice_coluna]) if len(linha) > indice_coluna else ""
            campos[str(identificador)] = valor

        nome_limpo = limpar_nome_arquivo(nome_arquivo)

        arquivo_docx_saida = os.path.join(estrutura["word"], f"{nome_limpo}.docx")
        arquivo_docx_saida = gerar_nome_unico(arquivo_docx_saida)

        preencher_content_controls(arquivo_word_base, arquivo_docx_saida, campos)
        total_word += 1

        if gerar_pdf:
            nome_pdf = os.path.splitext(os.path.basename(arquivo_docx_saida))[0] + ".pdf"
            arquivo_pdf_saida = os.path.join(estrutura["pdf"], nome_pdf)
            arquivo_pdf_saida = gerar_nome_unico(arquivo_pdf_saida)
            convert(arquivo_docx_saida, arquivo_pdf_saida)
            total_pdf += 1

    return {
        "arquivo_excel": arquivo_excel,
        "arquivo_word": arquivo_word_base,
        "saida": estrutura["saida"],
        "total_word": total_word,
        "total_pdf": total_pdf,
    }


def executar_conversao_word_para_pdf(pasta_raiz):
    estrutura = criar_estrutura_pastas_conversao(pasta_raiz)

    arquivos_word = obter_todos_arquivos(
        estrutura["base"],
        extensoes=[".docx"]
    )

    total_pdf = 0

    for arquivo_word in arquivos_word:
        nome_pdf = os.path.splitext(os.path.basename(arquivo_word))[0] + ".pdf"
        arquivo_pdf_saida = os.path.join(estrutura["pdf"], nome_pdf)
        arquivo_pdf_saida = gerar_nome_unico(arquivo_pdf_saida)

        convert(arquivo_word, arquivo_pdf_saida)
        total_pdf += 1

    return {
        "total_word_entrada": len(arquivos_word),
        "total_pdf": total_pdf,
        "saida": estrutura["saida"],
        "pdf": estrutura["pdf"],
    }